import csv
import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
import xlwings as xw
import tkinter as tk
from tkinter import filedialog
# install: pip install openpyxl, xlsxwriter, pandas
import re

def sanitize_sheet_name(name):
    # 1. Replace invalid characters with an underscore
    # Invalid: \ / * ? : [ ]
    clean_name = re.sub(r'[\\/*?:\[\]]', '_', name)
    
    # 2. Excel sheet names are limited to 31 characters
    return clean_name[:31]



def save_data_to_csv(data, headers=None, filename="output.csv"):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="Select Folder to Save CSV")
    
    if not folder_path:
        print("No folder selected.")
        return

    output_path = os.path.join(folder_path, filename)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        writer.writerows(data)
    print(f"Saved to {output_path}")       

def save_datum_to_csv(datum, filename="output", extender=".csv", headers=None ):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="Select Folder to Save CSV")

    if not folder_path:
        print("No folder selected.")
        return

 
    for i, file_data in enumerate(datum, start=1):
        save_data_to_csv_wo_dialog(
            file_data,
            headers=headers,
            folder_path=folder_path,
            filename=f"{filename}_{i}{extender}"
        )

def save_data_to_csv_wo_dialog(data, headers=None, folder_path=None, filename="output.csv"):
    # default: current folder
    if folder_path is None:
        folder_path = os.getcwd()

    output_path = os.path.join(folder_path, filename)
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if headers:
            writer.writerow(headers)
        writer.writerows(data)
    print(f"Saved to {output_path}")       

def save_datum_to_sp(datum, filename="output", headers=None ):
    root = tk.Tk()
    root.withdraw()
    folder_path = filedialog.askdirectory(title="Select Folder to Save CSV")

    if not folder_path:
        print("No folder selected.")
        return

 
    for i, file_data in enumerate(datum, start=1):
        save_data_to_csv_wo_dialog(
            file_data,
            headers=headers,
            folder_path=folder_path,
            filename=f"{filename}_{i}.sp"
        )




def load_data_from_csv(file_path=None):
    # 1. Check if a path was provided; if not, open the explorer
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="Select CSV file to load",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        # Handle user canceling the dialog
        if not file_path:
            print("No file selected.")
            return None, None

    # 2. Process the file (whether provided or selected)
    data = []
    headers = None

    try:
        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            
            # Extract headers from the first row
            headers = next(reader, None) 
            
            # Extract remaining rows
            data = [row for row in reader]
                
        print(f"Successfully loaded {len(data)} rows from: {file_path}")
        return data, headers

    except FileNotFoundError:
        print(f"Error: The file at {file_path} was not found.")
        return None, None
    except Exception as e:
        print(f"An error occurred: {e}")
        return None, None

def load_pandas_data_from_csv(file_path=None):
    # 1. Handle the path logic
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.askopenfilename(
            title="Select CSV file to load",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        # Destroy root to prevent hanging threads
        root.destroy()
        
        if not file_path:
            print("No file selected.")
            return None

    # 2. Use Pandas to load the data
    try:
        # Pandas automatically handles headers (row 0) and data types
        df = pd.read_csv(file_path, encoding="utf-8")
        
        print(f"Loaded {len(df)} rows from {file_path}")
        return df

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def load_all_csvs_to_list(extension='.csv', root_path=None):
    # 1. Handle the path logic (Selecting a FOLDER instead of a FILE)
    if not root_path:
        root = tk.Tk()
        root.withdraw()
        root_path = filedialog.askdirectory(title="Select Root Folder to Search for CSVs")
        root.destroy()
        
        if not root_path:
            print("No folder selected.")
            return []

    all_dataframes = []

    # 2. Walk through the directory tree
    # os.walk yields: root directory, subdirectories, and filenames
    for current_root, dirs, files in os.walk(root_path):
        for file in files:
            if file.lower().endswith(extension):
                file_full_path = os.path.join(current_root, file)
                
                try:
                    # Load the individual CSV
                    df = pd.read_csv(file_full_path, encoding="utf-8")
                    
                    # Store metadata (optional but helpful)
                    # df['source_file'] = file 
                    
                    all_dataframes.append(df)
                    print(f"Loaded: {file_full_path} ({len(df)} rows)")
                    
                except Exception as e:
                    print(f"Skipped {file} due to error: {e}")

    print(f"\nSuccessfully loaded {len(all_dataframes)} total DataFrames.")
    return all_dataframes

def load_all_csvs_to_dict(extension='.csv', root_path=None, use_file_path_as_name=False):
    # 1. Handle the path logic
    if not root_path:
        root = tk.Tk()
        root.withdraw()
        root_path = filedialog.askdirectory(title="Select Root Folder")
        root.destroy()
        
        if not root_path:
            print("No folder selected.")
            return {}

    # Initialize dictionary: { "filename": dataframe }
    dfs_dict = {}

    i = 0 
    # 2. Walk through the directory tree
    for current_root, dirs, files in os.walk(root_path):
        i += 1
        j = 0
        for file in files:
            j += 1
            if file.lower().endswith(extension):
                file_full_path = os.path.join(current_root, file)
                if use_file_path_as_name:
                    rel_path = sanitize_sheet_name(os.path.relpath(file_full_path, root_path)).replace(" ","")
                else:
                    rel_path = f"f{i}_{j}"
                
                try:
                    df = pd.read_csv(file_full_path, encoding="utf-8")
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')] 
                    # Store in dictionary using the relative path as the key
                    dfs_dict[rel_path] = df
                    print(f"Mapped '{rel_path}' with {len(df)} rows")
                    
                except Exception as e:
                    print(f"Error loading {file}: {e}")

    print(f"\nTotal files loaded into dictionary: {len(dfs_dict)}")
    return dfs_dict

def load_all_csvs_to_dict_filter(filter_list, extension='.csv', root_path=None, use_file_path_as_name=False):
    # 1. Handle the path logic
    if not root_path:
        root = tk.Tk()
        root.withdraw()
        root_path = filedialog.askdirectory(title="Select Root Folder")
        root.destroy()
        
        if not root_path:
            print("No folder selected.")
            return {}

    # Initialize dictionary: { "filename": dataframe }
    dfs_dict = {}

    i = 0 
    # 2. Walk through the directory tree
    for current_root, dirs, files in os.walk(root_path):
        i += 1
        j = 0
        for file in files:
            j += 1
            # Check if file has the correct extension
            if file.lower().endswith(extension):
                
                # --- FILTER LOGIC ---
                # Check if ANY word in the filter_list is present in the filename (case-insensitive)
                if not any(word.lower() in file.lower() for word in filter_list):
                    continue  # Skip this file if it doesn't match the filter
                # ---------------------

                file_full_path = os.path.join(current_root, file)
                if use_file_path_as_name:
                    # Note: Ensure sanitize_sheet_name is defined elsewhere in your script
                    rel_path = sanitize_sheet_name(os.path.relpath(file_full_path, root_path)).replace(" ", "")
                else:
                    rel_path = f"f{i}_{j}"
                
                try:
                    df = pd.read_csv(file_full_path, encoding="utf-8")
                    df = df.loc[:, ~df.columns.str.contains('^Unnamed')] 
                    
                    # Store in dictionary using the relative path as the key
                    dfs_dict[rel_path] = df
                    print(f"Mapped '{rel_path}' ({file}) with {len(df)} rows")
                    
                except Exception as e:
                    print(f"Error loading {file}: {e}")

    print(f"\nTotal files loaded into dictionary: {len(dfs_dict)}")
    return dfs_dict



def export_pandas_data_to_csv(df, file_path=None):
    # 1. Handle the path logic
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        # Use asksaveasfilename instead of askopenfilename
        file_path = filedialog.asksaveasfilename(
            title="Select location to save CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        # Destroy root to prevent hanging threads
        root.destroy()
        
        if not file_path:
            print("Export cancelled: No file path selected.")
            return False

    # 2. Use Pandas to export the data
    try:
        # index=False is usually preferred so you don't save the row numbers
        df.to_csv(file_path, index=False, encoding="utf-8")
        
        print(f"Successfully exported {len(df)} rows to {file_path}")
        return True

    except Exception as e:
        print(f"An error occurred during export: {e}")
        return False

def export_multiple_dfs_to_excel(df_dict, header_info=True, file_path=None):
    """
    df_dict: A dictionary like {"Sheet1": df1, "Sheet2": df2}
    """
    # 1. Handle the path logic
    if not file_path:
        root = tk.Tk()
        root.withdraw()
        file_path = filedialog.asksaveasfilename(
            title="Select location to save Excel Workbook",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        root.destroy()
        
        if not file_path:
            print("Export cancelled: No file path selected.")
            return False

    # 2. Use ExcelWriter to save multiple sheets
    try:
        # Use context manager (with) to ensure the file is closed properly
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            for sheet_name, df in df_dict.items():
                # Clean sheet name (Excel limited to 31 chars)
                clean_name = str(sheet_name)[:31]
                df.to_excel(writer, sheet_name=clean_name, index=False, header=header_info)
        
        print(f"Successfully exported {len(df_dict)} sheets to {file_path}")
        return True

    except Exception as e:
        print(f"An error occurred during Excel export: {e}")
        return False

def replace_xlsx_sheet_with_df(file_path, sheet_name, df, header_info=True):
    safe_sheet_name = sheet_name[:31]
    is_multi_sheet = False
    
    # 1. Attempt to evaluate the existing file safely
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            # If it has other sheets, we intend to append/preserve them
            if len(wb.sheetnames) > 1:
                is_multi_sheet = True
                if safe_sheet_name in wb.sheetnames:
                    del wb[safe_sheet_name]
                    wb.save(file_path)
            wb.close()
        except Exception:
            # Handles 'File is not a zip file', 0-byte sizes, or hard corruptions
            print(f"Warning: '{file_path}' is corrupted or unreadable. Re-creating file.")
            try:
                os.remove(file_path) # Clear out the corrupted file completely
            except OSError:
                pass
            is_multi_sheet = False

    try:
        # 2. Determine writing strategy based on file state
        write_mode = 'a' if is_multi_sheet else 'w'
        if_exists = 'overlay' if is_multi_sheet else None

        with pd.ExcelWriter(file_path, engine='openpyxl', mode=write_mode, if_sheet_exists=if_exists) as writer:
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False, header=header_info)
        
        print(f"Sheet '{safe_sheet_name}' successfully replaced/created in {file_path}")
        return True

    except Exception as e:
        print(f"An error occurred while writing the sheet: {e}")
        return False

def read_xlsx_sheet_to_df(file_path, sheet_name):
    try:
        df = pd.read_excel(
            file_path, 
            sheet_name=sheet_name,
            engine='openpyxl',
            engine_kwargs={'data_only': True}
            )
        print(f"Successfully read sheet '{sheet_name}' from {file_path}")
        return df
    except Exception as e:
        print(f"An error occurred while reading the sheet: {e}")
        return None

def read_xlsx_sheet_to_df2(file_path, sheet_name):
    app = xw.App(visible=False)
    wb = app.books.open(file_path)
    
    wb.app.calculate()  # force recalculation
    
    sht = wb.sheets[sheet_name]
    data = sht.used_range.value
    
    wb.close()
    app.quit()
    
    return pd.DataFrame(data[1:], columns=data[0])

# read_xlsx_sheet_to_df2 = formula evaluation save possible while read_xlsx_sheet_to_df returns none for formula cells.

def replace_first_rows(df, new_rows):
    if not isinstance(new_rows, pd.DataFrame):
        new_rows = pd.DataFrame(new_rows)

    n_rows, n_cols = new_rows.shape

    # Ensure df has enough columns
    if df.shape[1] < n_cols:
        for i in range(df.shape[1], n_cols):
            df[i] = ""

    # 🔥 Case 1: df is longer → remove top part
    if len(df) >= n_rows:
        df_tail = df.iloc[n_rows:].reset_index(drop=True)
    else:
        # 🔥 Case 2: df is shorter → just empty tail
        df_tail = pd.DataFrame(columns=df.columns)

    # Align columns
    new_rows.columns = range(n_cols)
    df_tail.columns = range(df.shape[1])

    # Combine cleanly
    result = pd.concat([new_rows, df_tail], ignore_index=True)

    return result




def export_sheet_to_csv_no_df(xlsx_path, sheet_name, output_csv_path):
    # load_workbook with data_only=True targets the cached data values instead of formulas
    wb = load_workbook(xlsx_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        return
        
    sheet = wb[sheet_name]
    
    # Open a clean CSV file to write to natively
    with open(output_csv_path, mode='w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file)
        
        # Loop row by row, grab cell values, and drop them into the CSV
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)
            
    print(f"Successfully exported {sheet_name} to {output_csv_path} without DataFrames!")


def export_cross_sheet_to_csv(xlsx_path, target_sheet_name, output_csv_path):
    # 1. Open Excel hidden in the background
    app = xw.App(visible=False)
    
    try:
        # 2. Open workbook (Excel automatically calculates all cross-sheet formulas)
        wb = xw.Book(xlsx_path)
        sheet = wb.sheets[target_sheet_name]
        
        # 3. Create a temporary, brand-new workbook to hold just the target sheet
        new_wb = xw.Book()
        sheet.copy(before=new_wb.sheets[0])
        
        # 4. Save this specific sheet directly as a clean CSV file
        # (Excel takes care of converting formulas to pure values for us!)
        new_wb.save(output_csv_path)
        new_wb.close()
        
        print(f"Successfully evaluated cross-sheet formulas and saved to {output_csv_path}")
        
    except Exception as e:
        print(f"An error occurred: {e}")
        
    finally:
        # 5. Clean up and close Excel safely no matter what happens
        try: wb.close() 
        except: pass
        app.quit()

# Example usage:
# export_sheet_to_csv_no_df("a.xlsx", "LM_cau", "ab_a.csv")
